from pathlib import Path

from openpyxl import load_workbook

from database import get_connection


def _to_float(value, default=0.0):
    if value is None:
        return default

    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip().replace(" ", "").replace(",", ".")
    if not text:
        return default

    try:
        return float(text)
    except ValueError:
        return default


def _norm(value):
    if value is None:
        return ""
    return str(value).strip()


def ensure_price_catalog_schema():
    with get_connection() as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS price_catalog (
                part_no TEXT PRIMARY KEY,
                check_digit TEXT,
                description TEXT,
                price_excl_vat REAL DEFAULT 0,
                discount_code TEXT,
                unit TEXT,
                product_group TEXT,
                function_group TEXT,
                statistic_no TEXT,
                weight REAL DEFAULT 0,
                country_of_origin TEXT,
                source_file TEXT,
                updated_at TEXT DEFAULT CURRENT_TIMESTAMP
            )
            """
        )

        conn.execute("CREATE INDEX IF NOT EXISTS idx_price_catalog_description ON price_catalog(description)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_price_catalog_discount_code ON price_catalog(discount_code)")

        conn.commit()


def import_dsp_price_file(excel_path):
    """
    Import rapide du fichier DSP price.

    Colonnes attendues :
    A = Part No
    B = Check digit
    C = Description
    D = Price excl VAT
    E = Discount Code
    F = Unit
    G = Product group
    H = Function group
    I = Statistic No
    J = Weight
    K = Country of Origin

    Optimisation :
    - lecture Excel en iter_rows(values_only=True)
    - préparation en mémoire
    - executemany SQLite en une seule transaction
    """
    ensure_price_catalog_schema()

    excel_path = Path(excel_path)
    if not excel_path.exists():
        raise FileNotFoundError(f"Fichier introuvable : {excel_path}")

    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = {}

    for index, value in enumerate(first_row):
        key = _norm(value).lower()
        if key:
            headers[key] = index

    required = ["part no", "description", "price excl vat", "discount code"]
    missing = [key for key in required if key not in headers]

    if missing:
        raise RuntimeError("Colonnes introuvables dans le fichier prix : " + ", ".join(missing))

    def idx(name):
        return headers.get(name)

    i_part = idx("part no")
    i_check = idx("check digit")
    i_desc = idx("description")
    i_price = idx("price excl vat")
    i_dc = idx("discount code")
    i_unit = idx("unit")
    i_product_group = idx("product group")
    i_function_group = idx("function group")
    i_stat = idx("statistic no")
    i_weight = idx("weight")
    i_origin = idx("country of origin")

    def get(row, i, default=""):
        if i is None:
            return default
        if i >= len(row):
            return default
        return row[i]

    source_file = excel_path.name
    records = []
    skipped = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        part_no = _norm(get(row, i_part))

        if not part_no:
            skipped += 1
            continue

        records.append(
            (
                part_no,
                _norm(get(row, i_check)),
                _norm(get(row, i_desc)),
                _to_float(get(row, i_price), 0.0),
                _norm(get(row, i_dc)),
                _norm(get(row, i_unit)),
                _norm(get(row, i_product_group)),
                _norm(get(row, i_function_group)),
                _norm(get(row, i_stat)),
                _to_float(get(row, i_weight), 0.0),
                _norm(get(row, i_origin)),
                source_file,
            )
        )

    with get_connection() as conn:
        conn.execute("PRAGMA synchronous = OFF")
        conn.execute("PRAGMA journal_mode = MEMORY")

        conn.executemany(
            """
            INSERT INTO price_catalog (
                part_no,
                check_digit,
                description,
                price_excl_vat,
                discount_code,
                unit,
                product_group,
                function_group,
                statistic_no,
                weight,
                country_of_origin,
                source_file,
                updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
            ON CONFLICT(part_no) DO UPDATE SET
                check_digit = excluded.check_digit,
                description = excluded.description,
                price_excl_vat = excluded.price_excl_vat,
                discount_code = excluded.discount_code,
                unit = excluded.unit,
                product_group = excluded.product_group,
                function_group = excluded.function_group,
                statistic_no = excluded.statistic_no,
                weight = excluded.weight,
                country_of_origin = excluded.country_of_origin,
                source_file = excluded.source_file,
                updated_at = CURRENT_TIMESTAMP
            """,
            records,
        )

        conn.commit()

    return {
        "source_file": source_file,
        "imported": len(records),
        "skipped": skipped,
    }


def lookup_part(part_no):
    ensure_price_catalog_schema()

    reference = _norm(part_no)
    if not reference:
        return None

    with get_connection() as conn:
        row = conn.execute(
            """
            SELECT *
            FROM price_catalog
            WHERE part_no = ?
            """,
            (reference,),
        ).fetchone()

    return row


def get_catalog_count():
    ensure_price_catalog_schema()

    with get_connection() as conn:
        count = conn.execute("SELECT COUNT(*) FROM price_catalog").fetchone()[0]
        last = conn.execute(
            """
            SELECT source_file, updated_at
            FROM price_catalog
            WHERE source_file IS NOT NULL
            ORDER BY updated_at DESC
            LIMIT 1
            """
        ).fetchone()

    return {
        "count": count,
        "source_file": last["source_file"] if last else "",
        "updated_at": last["updated_at"] if last else "",
    }


def search_catalog(query, limit=30):
    ensure_price_catalog_schema()

    q = f"%{_norm(query)}%"

    with get_connection() as conn:
        rows = conn.execute(
            """
            SELECT part_no, description, price_excl_vat, discount_code
            FROM price_catalog
            WHERE part_no LIKE ?
               OR description LIKE ?
            ORDER BY part_no
            LIMIT ?
            """,
            (q, q, int(limit)),
        ).fetchall()

    return rows
