"""
Importeur ServiceCalculationExport.xlsx

But : lire un export Service Calculator et produire un JSON propre,
utilisable ensuite pour créer un devis brouillon dans le futur soft.

Usage :
    python service_calculation_importer.py /chemin/ServiceCalculationExport.xlsx --out summary.json

Dépendance :
    pip install openpyxl
"""

from __future__ import annotations

import argparse
import json
import re
from dataclasses import dataclass, asdict
from datetime import datetime, date
from pathlib import Path
from typing import Any, Optional

from openpyxl import load_workbook


INTERVENTION_SHEET_RE = re.compile(r"^(\d{4}-\d{2}-\d{2})\s*\((\d+)\)$")
SERVICE_GROUP_RE = re.compile(r"^[A-Z]$")


@dataclass
class EngineInfo:
    unit: Optional[str] = None
    installation: Optional[str] = None
    chassis_id: Optional[str] = None
    serial_number: Optional[str] = None
    product_category: Optional[str] = None
    product_name: Optional[str] = None
    product_designation: Optional[str] = None
    product_part_number: Optional[str] = None
    status: Optional[str] = None
    current_country: Optional[str] = None


@dataclass
class CalculationBasis:
    price_list_used: Optional[str] = None
    currency: Optional[str] = None
    total_calculation_hours: Optional[float] = None
    op_hours_per_year: Optional[float] = None
    labour_rate: Optional[float] = None
    number_of_service_interventions: Optional[int] = None


@dataclass
class CalculationResult:
    cost_per_hour: Optional[float] = None
    total_labour_hours: Optional[float] = None
    total_labour_cost: Optional[float] = None
    total_material_cost: Optional[float] = None
    total_additional_material_cost: Optional[float] = None
    misc_cost: Optional[float] = None
    total: Optional[float] = None


def clean(value: Any) -> Any:
    """Nettoie les valeurs Excel pour les rendre JSON-friendly."""
    if isinstance(value, str):
        value = value.strip()
        return value if value else None
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return value


def to_float(value: Any) -> Optional[float]:
    value = clean(value)
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def to_int(value: Any) -> Optional[int]:
    number = to_float(value)
    if number is None:
        return None
    return int(number)


def normalize_key(label: str) -> str:
    return (
        label.strip()
        .lower()
        .replace("/", "_")
        .replace(" ", "_")
        .replace("-", "_")
        .replace("__", "_")
    )



# --- Multilingual Service Calculator import helpers ---
def _normalize_import_text(value):
    import unicodedata
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = text.replace("œ", "oe").replace("’", "'").replace("`", "'")
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = " ".join(text.split())
    return text


SHEET_ALIASES = {
    "First page": [
        "First page",
        "Première page",
        "Premiere page",
        "Page de garde",
        "Page initiale",
        "Accueil",
    ],
    "Hidden for import": [
        "Hidden for import",
        "Caché pour l'importation",
        "Cache pour l'importation",
        "Caché pour importation",
        "Cache pour importation",
    ],
    "Overview": [
        "Overview",
        "Présentation",
        "Presentation",
        "Aperçu",
        "Apercu",
        "Synthèse",
        "Synthese",
        "Résumé",
        "Resume",
    ],
}


def _is_labour_label(value):
    return _normalize_import_text(value) in {
        "labour",
        "labor",
        "main d'oeuvre",
        "main d oeuvre",
        "main-doeuvre",
        "main-d'oeuvre",
    }
# --- End multilingual helpers ---



def get_ws(wb, name):
    """
    Retourne un onglet Excel en acceptant les noms anglais/français.
    """
    candidates = SHEET_ALIASES.get(name, [name])

    for candidate in candidates:
        if candidate in wb.sheetnames:
            return wb[candidate]

    normalized_sheets = {
        _normalize_import_text(sheet_name): sheet_name
        for sheet_name in wb.sheetnames
    }

    for candidate in candidates:
        normalized = _normalize_import_text(candidate)
        if normalized in normalized_sheets:
            return wb[normalized_sheets[normalized]]

    aliases = ", ".join(candidates)
    found = ", ".join(wb.sheetnames)
    raise ValueError(
        f"Onglet manquant : {name}. Alias acceptés : {aliases}. Onglets trouvés : {found}"
    )


def parse_first_page(wb):
    """
    Lit la première page du Service Calculator en anglais ou en français.

    Champs anglais :
    - First page
    - Serial Number, Product Name, Product Designation, Current Country
    - Total No of calculation hours, Op hrs / year, Labour rate...

    Champs français :
    - Première page
    - Numéro de série, Nom du produit, Désignation du produit, Pays actuel
    - Nombre total d'heures calculé, Heures/année de fonctionnement, Taux de main d'œuvre...
    """
    ws = get_ws(wb, "First page")

    def cell_value(addr):
        value = ws[addr].value
        if isinstance(value, str):
            value = value.replace("\xa0", " ").strip()
        return value

    def to_float(value):
        if value is None or value == "":
            return None
        if isinstance(value, (int, float)):
            return float(value)
        text = str(value).replace("\xa0", " ").replace(" ", "").replace(",", ".").strip()
        try:
            return float(text)
        except Exception:
            return None

    def infer_country(price_list, current_country):
        if current_country:
            return current_country

        text = str(price_list or "").lower()
        if "france" in text or "-fr" in text or "_fr" in text:
            return "FR"
        if "belgium" in text or "belgique" in text:
            return "BE"
        if "netherlands" in text or "pays-bas" in text:
            return "NL"
        if "germany" in text or "allemagne" in text:
            return "DE"
        if "spain" in text or "espagne" in text:
            return "ES"
        if "italy" in text or "italie" in text:
            return "IT"
        return None

    price_list_used = cell_value("A6")

    engine = {
        "unit": cell_value("A3"),
        "installation": cell_value("B3"),
        "chassis_id": cell_value("C3"),
        "serial_number": cell_value("D3"),
        "product_category": cell_value("E3"),
        "product_name": cell_value("F3"),
        "product_designation": cell_value("G3"),
        "product_part_number": cell_value("H3"),
        "status": cell_value("I3"),
        "current_country": infer_country(price_list_used, cell_value("J3")),
    }

    calculation_basis = {
        "price_list_used": price_list_used,
        "currency": cell_value("B8"),
        "total_calculation_hours": to_float(cell_value("B9")),
        "op_hours_per_year": to_float(cell_value("B10")),
        "labour_rate": to_float(cell_value("B11")),
        "number_of_service_interventions": int(to_float(cell_value("B12")) or 0),
    }

    calculation_result = {
        "cost_per_hour": to_float(cell_value("B15")),
        "total_labour_hours": to_float(cell_value("B16")),
        "total_labour_cost": to_float(cell_value("B17")),
        "total_material_cost": to_float(cell_value("B18")),
        "total_additional_material_cost": to_float(cell_value("B19")),
        "misc_cost": to_float(cell_value("B20")),
        "total": to_float(cell_value("B21")),
    }

    return {
        "engine": engine,
        "calculation_basis": calculation_basis,
        "calculation_result": calculation_result,
    }

def parse_hidden_for_import(wb) -> list[dict[str, Any]]:
    ws = get_ws(wb, "Hidden for import")
    lines: list[dict[str, Any]] = []
    current_group: Optional[str] = None

    for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        component, price, time, part_no, qty = [clean(v) for v in row[:5]]
        if component is None:
            continue

        component_text = str(component)
        is_group = bool(SERVICE_GROUP_RE.match(component_text)) and price is None and time is None and part_no is None

        if is_group:
            current_group = component_text
            line_type = "service_group"
        elif component_text.lower() in ("labour", "labor", "main d’œuvre", "main d'œuvre", "main d oeuvre", "main d'oeuvre"):
            line_type = "labour"
        elif part_no is not None:
            line_type = "part"
        else:
            line_type = "operation"

        lines.append(
            {
                "source_row": row_number,
                "group": current_group,
                "line_type": line_type,
                "component": component_text,
                "unit_price": to_float(price),
                "labour_time": to_float(time),
                "part_number": str(part_no) if part_no is not None else None,
                "quantity": to_float(qty),
            }
        )

    return lines


def intervention_dates_from_sheets(wb) -> dict[str, int]:
    dates: dict[str, int] = {}
    for name in wb.sheetnames:
        match = INTERVENTION_SHEET_RE.match(name)
        if match:
            dates[match.group(1)] = int(match.group(2))
    return dates


def parse_overview(wb) -> dict[str, Any]:
    ws = get_ws(wb, "Overview")
    sheet_date_hours = intervention_dates_from_sheets(wb)

    headers = [clean(cell.value) for cell in ws[1]]
    date_columns: list[tuple[int, str]] = []
    for index, header in enumerate(headers, start=1):
        if isinstance(header, str) and re.match(r"^\d{4}-\d{2}-\d{2}$", header):
            date_columns.append((index, header))

    lines: list[dict[str, Any]] = []
    totals_by_name: dict[str, dict[str, Any]] = {}
    current_group: Optional[str] = None

    for row_number in range(2, ws.max_row + 1):
        component = clean(ws.cell(row_number, 1).value)
        if component is None:
            continue
        component_text = str(component)

        line = {
            "source_row": row_number,
            "component": component_text,
            "price": to_float(ws.cell(row_number, 2).value),
            "total_cost": to_float(ws.cell(row_number, 3).value),
            "total_time": to_float(ws.cell(row_number, 4).value),
            "interval_hours": to_float(ws.cell(row_number, 5).value),
            "interval_months": to_float(ws.cell(row_number, 6).value),
            "part_number": str(clean(ws.cell(row_number, 7).value)) if clean(ws.cell(row_number, 7).value) is not None else None,
            "quantity": to_float(ws.cell(row_number, 8).value),
            "group": current_group,
            "schedule_quantities": {},
        }

        is_group = bool(SERVICE_GROUP_RE.match(component_text)) and line["price"] is None
        if is_group:
            current_group = component_text
            line["group"] = current_group
            line["line_type"] = "service_group"
        elif component_text in {"Parts", "Labour", "Misc", "Total", "Accumulated"}:
            line["line_type"] = "summary_total"
            totals_by_name[component_text.lower()] = line
        elif component_text.lower() in ("labour", "labor", "main d’œuvre", "main d'œuvre", "main d oeuvre", "main d'oeuvre"):
            line["line_type"] = "labour"
        elif line["part_number"]:
            line["line_type"] = "part"
        else:
            line["line_type"] = "operation"

        for col_index, date_str in date_columns:
            qty = to_float(ws.cell(row_number, col_index).value)
            if qty is not None:
                line["schedule_quantities"][date_str] = qty

        lines.append(line)

    interventions: list[dict[str, Any]] = []
    for col_index, date_str in date_columns:
        interventions.append(
            {
                "date": date_str,
                "engine_hours": sheet_date_hours.get(date_str),
                "parts_cost": to_float(ws.cell(totals_by_name["parts"]["source_row"], col_index).value) if "parts" in totals_by_name else None,
                "labour_cost": to_float(ws.cell(totals_by_name["labour"]["source_row"], col_index).value) if "labour" in totals_by_name else None,
                "misc_cost": to_float(ws.cell(totals_by_name["misc"]["source_row"], col_index).value) if "misc" in totals_by_name else None,
                "total_cost": to_float(ws.cell(totals_by_name["total"]["source_row"], col_index).value) if "total" in totals_by_name else None,
                "accumulated_cost": to_float(ws.cell(totals_by_name["accumulated"]["source_row"], col_index).value) if "accumulated" in totals_by_name else None,
            }
        )

    return {
        "overview_lines": lines,
        "interventions": interventions,
    }


def parse_intervention_sheets(wb) -> list[dict[str, Any]]:
    interventions: list[dict[str, Any]] = []

    for sheet_name in wb.sheetnames:
        match = INTERVENTION_SHEET_RE.match(sheet_name)
        if not match:
            continue

        ws = wb[sheet_name]
        current_group: Optional[str] = None
        lines: list[dict[str, Any]] = []

        for row_number in range(2, ws.max_row + 1):
            component = clean(ws.cell(row_number, 1).value)
            if component is None:
                continue
            component_text = str(component)
            price = to_float(ws.cell(row_number, 2).value)
            part_no = clean(ws.cell(row_number, 7).value)

            is_group = bool(SERVICE_GROUP_RE.match(component_text)) and price is None
            if is_group:
                current_group = component_text
                line_type = "service_group"
            elif component_text.lower() in ("labour", "labor", "main d’œuvre", "main d'œuvre", "main d oeuvre", "main d'oeuvre"):
                line_type = "labour"
            elif part_no is not None:
                line_type = "part"
            else:
                line_type = "operation"

            lines.append(
                {
                    "source_row": row_number,
                    "group": current_group,
                    "line_type": line_type,
                    "component": component_text,
                    "price": price,
                    "total_cost": to_float(ws.cell(row_number, 3).value),
                    "time": to_float(ws.cell(row_number, 4).value),
                    "interval_hours": to_float(ws.cell(row_number, 5).value),
                    "interval_months": to_float(ws.cell(row_number, 6).value),
                    "part_number": str(part_no) if part_no is not None else None,
                    "quantity": to_float(ws.cell(row_number, 8).value),
                }
            )

        interventions.append(
            {
                "sheet_name": sheet_name,
                "date": match.group(1),
                "engine_hours": int(match.group(2)),
                "lines": lines,
            }
        )

    return interventions


def parse_profitability(wb) -> dict[str, Any]:
    if "Profitabillity" not in wb.sheetnames:
        return {}

    ws = wb["Profitabillity"]
    rows: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        label = clean(row[0])
        if label is None:
            continue
        rows.append(
            {
                "category": str(label),
                "customer_cost": to_float(row[1]),
                "workshop_cost": to_float(row[2]),
                "profitability": to_float(row[3]),
            }
        )
    return {"rows": rows}


def build_quote_draft(summary: dict[str, Any]) -> dict[str, Any]:
    """Crée un premier brouillon de devis exploitable par la future application."""
    engine = summary["engine"]
    basis = summary["calculation_basis"]
    result = summary["calculation_result"]

    total_hours = basis.get("total_calculation_hours")
    op_hours_per_year = basis.get("op_hours_per_year")
    duration_years = None
    if total_hours and op_hours_per_year:
        duration_years = total_hours / op_hours_per_year

    return {
        "status": "draft_imported",
        "client": None,
        "engine": engine,
        "contract": {
            "currency": basis.get("currency"),
            "duration_years": duration_years,
            "total_hours": total_hours,
            "hours_per_year": op_hours_per_year,
            "number_of_interventions": basis.get("number_of_service_interventions"),
        },
        "base_costs": {
            "parts": result.get("total_material_cost"),
            "labour": result.get("total_labour_cost"),
            "misc": result.get("misc_cost"),
            "total": result.get("total"),
            "cost_per_hour": result.get("cost_per_hour"),
        },
        "dealer_parameters_to_add_later": {
            "parts_margin_percent": None,
            "labour_margin_percent": None,
            "admin_fee_percent": None,
            "logistics_fee_percent": None,
            "travel_cost": None,
            "indexation_percent_per_year": None,
        },
    }


def import_service_calculation(path: str | Path) -> dict[str, Any]:
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(path)

    wb = load_workbook(path, data_only=True, read_only=True)

    summary: dict[str, Any] = {
        "source_file": path.name,
        "import_type": "ServiceCalculationExport",
        "imported_at": datetime.now().isoformat(timespec="seconds"),
        "workbook_sheets": wb.sheetnames,
    }
    summary.update(parse_first_page(wb))
    summary["hidden_import_lines"] = parse_hidden_for_import(wb)
    overview = parse_overview(wb)
    summary.update(overview)
    summary["intervention_details"] = parse_intervention_sheets(wb)
    summary["profitability"] = parse_profitability(wb)
    summary["quote_draft"] = build_quote_draft(summary)

    return summary


def main() -> None:
    parser = argparse.ArgumentParser(description="Importe un ServiceCalculationExport.xlsx en JSON.")
    parser.add_argument("input", help="Chemin vers ServiceCalculationExport.xlsx")
    parser.add_argument("--out", help="Fichier JSON de sortie")
    parser.add_argument("--pretty", action="store_true", help="Affiche un JSON indenté")
    args = parser.parse_args()

    data = import_service_calculation(args.input)
    json_text = json.dumps(data, ensure_ascii=False, indent=2 if args.pretty else None)

    if args.out:
        Path(args.out).write_text(json_text, encoding="utf-8")
        print(f"Import OK - {args.out}")
        result = data["calculation_result"]
        engine = data["engine"]
        print(f"Moteur : {engine.get('product_designation')} / SN {engine.get('serial_number')}")
        print(f"Total : {result.get('total')} {data['calculation_basis'].get('currency')}")
        print(f"Interventions : {len(data.get('interventions', []))}")
    else:
        print(json_text)


if __name__ == "__main__":
    main()
