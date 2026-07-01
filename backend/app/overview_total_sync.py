from pathlib import Path

from openpyxl import load_workbook

from database import get_connection


def _to_number(value):
    if value is None:
        return None

    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip().replace(" ", "").replace(",", ".")
    if not text:
        return None

    try:
        return float(text)
    except ValueError:
        return None


def read_overview_column_c_totals(excel_path):
    excel_path = Path(excel_path)

    if not excel_path.exists():
        raise FileNotFoundError(f"Fichier Excel introuvable : {excel_path}")

    workbook = load_workbook(excel_path, data_only=True, read_only=True)

    if "Overview" not in workbook.sheetnames:
        raise RuntimeError("Onglet Overview introuvable dans le fichier Excel.")

    sheet = workbook["Overview"]
    found = {}

    for row in range(1, sheet.max_row + 1):
        label = sheet.cell(row=row, column=1).value
        value = sheet.cell(row=row, column=3).value

        if label is None:
            continue

        key = str(label).strip().lower()
        numeric_value = _to_number(value)

        if key in ("parts", "labour", "labor", "misc", "total"):
            found[key] = numeric_value

    parts = found.get("parts")
    labour = found.get("labour", found.get("labor"))
    misc = found.get("misc")
    total = found.get("total")

    if total is None:
        raise RuntimeError("Total Overview colonne C introuvable.")

    return {
        "parts": parts or 0.0,
        "labour": labour or 0.0,
        "misc": misc or 0.0,
        "total": total,
    }


def apply_overview_total_to_service_2_2(quote_id, excel_path):
    totals = read_overview_column_c_totals(excel_path)
    total = float(totals["total"])

    with get_connection() as conn:
        existing = conn.execute(
            """
            SELECT id
            FROM quote_services
            WHERE quote_id = ? AND service_id = '2,2'
            """,
            (quote_id,),
        ).fetchone()

        if existing is None:
            conn.execute(
                """
                INSERT INTO quote_services (
                    quote_id,
                    service_id,
                    service_group,
                    service_name,
                    source_excel,
                    included,
                    work_time_hours,
                    quantity,
                    unit_price,
                    fixed_price,
                    extra_travel,
                    calculated_price,
                    notes
                )
                VALUES (?, '2,2', 'Maintenance & repair', 'Maintenance parts & labour',
                        'Overview column C total', 0, 0, 0, 0, ?, 'Exclude', ?,
                        'Auto from Overview column C')
                """,
                (quote_id, total, total),
            )
        else:
            conn.execute(
                """
                UPDATE quote_services
                SET
                    work_time_hours = 0,
                    quantity = 0,
                    unit_price = 0,
                    fixed_price = ?,
                    extra_travel = 'Exclude',
                    calculated_price = ?,
                    source_excel = 'Overview column C total',
                    notes = ?
                WHERE quote_id = ? AND service_id = '2,2'
                """,
                (
                    total,
                    total,
                    f"Auto from Overview column C. Parts={totals['parts']}, Labour={totals['labour']}, Misc={totals['misc']}, Total={totals['total']}",
                    quote_id,
                ),
            )

        conn.commit()

    return totals
