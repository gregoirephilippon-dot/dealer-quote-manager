
from __future__ import annotations

import sqlite3
import unicodedata
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def _normalize(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = text.replace("œ", "oe").replace("’", "'").replace("`", "'")
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.replace("°", "")
    text = " ".join(text.split())
    return text


def _to_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace("\xa0", " ").replace(" ", "").replace(",", ".").strip()
    text = text.replace("EUR", "").replace("€", "")
    try:
        return float(text)
    except Exception:
        return None


def _get_sheet_any(wb, aliases: list[str]):
    normalized = {_normalize(name): name for name in wb.sheetnames}
    for alias in aliases:
        if _normalize(alias) in normalized:
            return wb[normalized[_normalize(alias)]]
    return None


def _get_first_page_totals(wb) -> dict[str, float]:
    ws = _get_sheet_any(wb, ["First page", "Première page", "Premiere page"])
    if ws is None:
        return {}

    return {
        "first_page_labour": _to_float(ws["B17"].value) or 0.0,
        "first_page_material": _to_float(ws["B18"].value) or 0.0,
        "first_page_additional_material": _to_float(ws["B19"].value) or 0.0,
        "first_page_misc": _to_float(ws["B20"].value) or 0.0,
        "first_page_total": _to_float(ws["B21"].value) or 0.0,
    }


def _find_header_columns(ws) -> dict[str, int]:
    aliases = {
        "component": {"component", "composant", "description"},
        "unit_price": {"price", "prix", "unit price", "prix unit", "prix unitaire", "prix unit."},
        "time_h": {"time", "duree", "durée", "temps", "temps h", "temps heure"},
        "part_no": {"part no", "part number", "n de ref", "n ref", "no de ref", "n° de ref", "n° de réf", "reference", "référence"},
        "qty": {"qty", "qte", "qté", "quantity", "quantite", "quantité"},
    }

    for row in range(1, min(ws.max_row, 20) + 1):
        found: dict[str, int] = {}
        for col in range(1, ws.max_column + 1):
            label = _normalize(ws.cell(row, col).value)
            for key, names in aliases.items():
                if label in {_normalize(x) for x in names}:
                    found[key] = col
        if {"component", "unit_price", "time_h", "part_no", "qty"}.issubset(found):
            found["_header_row"] = row
            return found

    # Fallback structure known for Hidden for import / Caché pour l'importation.
    return {
        "_header_row": 1,
        "component": 1,
        "unit_price": 2,
        "time_h": 3,
        "part_no": 4,
        "qty": 5,
    }


def _is_labour(component: Any) -> bool:
    label = _normalize(component)
    return label in {
        "labour",
        "labor",
        "main d'oeuvre",
        "main d oeuvre",
        "main-doeuvre",
        "main-d'oeuvre",
    }


def _is_additional_material(component: Any) -> bool:
    label = _normalize(component)
    return (
        "additional material" in label
        or "materiel supplementaire" in label
        or "materiau supplementaire" in label
        or "materiaux supplementaires" in label
    )


def calculate_service_2_2_from_excel(excel_path: str | Path) -> dict[str, Any]:
    wb = load_workbook(excel_path, read_only=True, data_only=True)

    try:
        ws = _get_sheet_any(
            wb,
            [
                "Hidden for import",
                "Caché pour l'importation",
                "Cache pour l'importation",
                "Caché pour importation",
                "Cache pour importation",
            ],
        )
        if ws is None:
            raise ValueError("Onglet Hidden for import / Caché pour l'importation introuvable")

        cols = _find_header_columns(ws)
        header_row = cols["_header_row"]

        labour_total = 0.0
        material_total = 0.0
        additional_material_total = 0.0
        current_material_block = 0.0
        detail_rows = 0
        labour_rows = 0
        material_rows = 0
        additional_rows = 0

        for row in range(header_row + 1, ws.max_row + 1):
            component = ws.cell(row, cols["component"]).value
            unit_price = _to_float(ws.cell(row, cols["unit_price"]).value)
            time_h = _to_float(ws.cell(row, cols["time_h"]).value)
            qty = _to_float(ws.cell(row, cols["qty"]).value)
            part_no = ws.cell(row, cols["part_no"]).value

            if component is None and unit_price is None and time_h is None and qty is None:
                continue

            if _is_labour(component):
                if unit_price is not None and time_h is not None:
                    labour_total += unit_price * time_h
                    labour_rows += 1
                    detail_rows += 1
                continue

            if _is_additional_material(component):
                # Dans les fichiers Volvo, la ligne contient un pourcentage,
                # par exemple 10 %, à appliquer au bloc de pièces précédent.
                percent = unit_price or 0.0
                add_value = current_material_block * percent / 100.0
                additional_material_total += add_value
                additional_rows += 1
                detail_rows += 1
                current_material_block = 0.0
                continue

            if unit_price is not None and qty is not None:
                value = unit_price * qty
                material_total += value
                current_material_block += value
                material_rows += 1
                detail_rows += 1

        first_page = _get_first_page_totals(wb)

        detail_total = labour_total + material_total + additional_material_total + first_page.get("first_page_misc", 0.0)
        first_total = first_page.get("first_page_total", 0.0)
        diff = detail_total - first_total if first_total else 0.0

        source_note = (
            "Source 2.2 détaillée: Hidden for import / Caché pour l'importation. "
            "Calcul: Main d'œuvre = Prix x Durée; Pièces = Prix x Qté; "
            "Matériel supplémentaire = % appliqué au bloc de pièces précédent. "
            f"Détail calculé = {detail_total:.3f}; "
            f"Main d'œuvre = {labour_total:.3f}; "
            f"Pièces = {material_total:.3f}; "
            f"Matériel supplémentaire = {additional_material_total:.3f}; "
            f"Divers = {first_page.get('first_page_misc', 0.0):.3f}; "
            f"Contrôle First page / Première page B21 = {first_total:.3f}; "
            f"Écart = {diff:.3f}."
        )

        return {
            "labour_total": labour_total,
            "material_total": material_total,
            "additional_material_total": additional_material_total,
            "misc_total": first_page.get("first_page_misc", 0.0),
            "detail_total": detail_total,
            "first_page_total": first_total,
            "difference": diff,
            "detail_rows": detail_rows,
            "labour_rows": labour_rows,
            "material_rows": material_rows,
            "additional_rows": additional_rows,
            "source_note": source_note,
        }
    finally:
        wb.close()


def _connect_db():
    here = Path(__file__).resolve()
    project_root = here.parents[2]
    db_path = project_root / "data" / "dealer_quote_manager.sqlite"
    return sqlite3.connect(db_path)


def apply_service_2_2_detail_calculation(quote_id: int, excel_path: str | Path) -> dict[str, Any]:
    result = calculate_service_2_2_from_excel(excel_path)

    # Si le calcul détaillé correspond au contrôle First page, on utilise le détail.
    # S'il y a un petit écart d'arrondi, le détail reste la source.
    total = result["detail_total"]

    conn = _connect_db()
    conn.row_factory = sqlite3.Row
    try:
        cols = {row["name"] for row in conn.execute("PRAGMA table_info(quote_services)").fetchall()}

        update_values = {
            "service_name": "Maintenance parts & labour",
            "service_group": "2. Maintenance",
            "work_time_hours": result["labour_total"],
            "quantity": result["material_rows"],
            "unit_price": 0.0,
            "fixed_price": total,
            "calculated_price": total,
            "source_excel": "Hidden for import detailed calculation",
            "notes": result["source_note"],
        }

        existing = conn.execute(
            "SELECT id FROM quote_services WHERE quote_id = ? AND service_id = ?",
            (quote_id, "2,2"),
        ).fetchone()

        if existing:
            assignments = []
            params = []
            for key, value in update_values.items():
                if key in cols:
                    assignments.append(f"{key} = ?")
                    params.append(value)
            params.append(quote_id)
            params.append("2,2")
            conn.execute(
                f"UPDATE quote_services SET {', '.join(assignments)} WHERE quote_id = ? AND service_id = ?",
                params,
            )
        else:
            insert_values = {
                "quote_id": quote_id,
                "service_id": "2,2",
                **update_values,
                "included": 1,
                "extra_travel": "No",
            }
            keys = [key for key in insert_values.keys() if key in cols]
            placeholders = ", ".join("?" for _ in keys)
            conn.execute(
                f"INSERT INTO quote_services ({', '.join(keys)}) VALUES ({placeholders})",
                [insert_values[key] for key in keys],
            )

        conn.commit()
        return result
    finally:
        conn.close()
